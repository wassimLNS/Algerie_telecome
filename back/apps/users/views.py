from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from django.utils import timezone

from .models import Utilisateur, HistoriqueConnexion, Role, LigneTelephonique, DemandeIT, StatutDemande
from .serializers import (
    LoginClientSerializer,
    LoginAgentSerializer,
    UtilisateurProfilSerializer,
    CreerAgentSerializer,
    ModifierAgentSerializer,
    AgentListSerializer,
    HistoriqueConnexionSerializer,
    LigneTelephoniqueSerializer,
    DemandeITSerializer,
    CreerDemandeSerializer,
)
from .permissions import EstAdmin, EstAdminIT, EstAdminOuAdminIT, EstClient, EstAgentOuPlus


def get_tokens_for_user(user):
    """Génère access token + refresh token pour un utilisateur"""
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


def log_connexion(utilisateur, request, succes=True, raison_echec=None):
    """Enregistre une tentative de connexion"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')

    HistoriqueConnexion.objects.create(
        utilisateur=utilisateur,
        ip_adresse=ip,
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        succes=succes,
        raison_echec=raison_echec,
    )


# ============================================================
# LOGIN CLIENT
# ============================================================
class LoginClientView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginClientSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']

            # Mettre à jour la dernière connexion
            user.derniere_connexion = timezone.now()
            user.save(update_fields=['derniere_connexion'])

            # Logger la connexion
            log_connexion(user, request, succes=True)

            tokens = get_tokens_for_user(user)
            return Response({
                'message': 'Connexion réussie',
                'user': UtilisateurProfilSerializer(user).data,
                'tokens': tokens,
            }, status=status.HTTP_200_OK)

        # Logger l'échec si on trouve l'utilisateur
        telephone = request.data.get('telephone')
        if telephone:
            try:
                user = Utilisateur.objects.get(telephone=telephone)
                log_connexion(user, request, succes=False, raison_echec='mot_de_passe_incorrect')
            except Utilisateur.DoesNotExist:
                pass

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# LOGIN AGENT / ADMIN
# ============================================================
class LoginAgentView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginAgentSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']

            # Mettre à jour la dernière connexion
            user.derniere_connexion = timezone.now()
            user.save(update_fields=['derniere_connexion'])

            # Logger la connexion
            log_connexion(user, request, succes=True)

            tokens = get_tokens_for_user(user)
            return Response({
                'message': 'Connexion réussie',
                'user': UtilisateurProfilSerializer(user).data,
                'tokens': tokens,
            }, status=status.HTTP_200_OK)

        # Logger l'échec
        email = request.data.get('email')
        if email:
            try:
                user = Utilisateur.objects.get(email=email)
                log_connexion(user, request, succes=False, raison_echec='mot_de_passe_incorrect')
            except Utilisateur.DoesNotExist:
                pass

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# LOGOUT
# ============================================================
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get('refresh')
            token = RefreshToken(refresh_token)
            token.blacklist()

            # Mettre à jour la déconnexion dans l'historique
            derniere = HistoriqueConnexion.objects.filter(
                utilisateur=request.user,
                succes=True,
                deconnecte_a__isnull=True
            ).last()
            if derniere:
                derniere.deconnecte_a = timezone.now()
                derniere.save()

            return Response({'message': 'Déconnexion réussie'}, status=status.HTTP_200_OK)
        except Exception:
            return Response({'error': 'Token invalide'}, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# MON PROFIL
# ============================================================
class MonProfilView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UtilisateurProfilSerializer(request.user)
        return Response(serializer.data)

    def put(self, request):
        serializer = UtilisateurProfilSerializer(
            request.user, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# GESTION DES AGENTS (admin lecture + audit, admin_it CRUD)
# ============================================================
class AgentsView(APIView):
    permission_classes = [IsAuthenticated]

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.role not in [Role.ADMIN, Role.ADMIN_IT]:
            self.permission_denied(request)

    def get(self, request):
        """Liste du staff — admin voit ses agents, admin_it voit TOUT (agents + managers)"""
        if request.user.role == Role.ADMIN:
            agents = Utilisateur.objects.filter(
                centre=request.user.centre,
                role__in=[Role.AGENT, Role.AGENT_TECHNIQUE, Role.AGENT_ANNEXE]
            )
        else:
            # Admin IT voit tout le staff (agents + managers)
            agents = Utilisateur.objects.filter(
                role__in=[Role.AGENT, Role.AGENT_TECHNIQUE, Role.AGENT_ANNEXE, Role.ADMIN]
            )
            # Filtres optionnels
            centre_id = request.query_params.get('centre')
            if centre_id:
                agents = agents.filter(centre_id=centre_id)
            role_filter = request.query_params.get('role')
            if role_filter:
                agents = agents.filter(role=role_filter)
        serializer = AgentListSerializer(agents, many=True)
        return Response(serializer.data)

    def post(self, request):
        """Créer un nouvel agent (admin_it uniquement)"""
        if request.user.role != Role.ADMIN_IT:
            return Response({'error': 'Seul l\'admin IT peut créer des agents'}, status=status.HTTP_403_FORBIDDEN)
        serializer = CreerAgentSerializer(data=request.data)
        if serializer.is_valid():
            agent = serializer.save()
            return Response(
                UtilisateurProfilSerializer(agent).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ClientsView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        """Liste des clients (filtrés par centre si l'admin appartient à un centre)"""
        clients = Utilisateur.objects.filter(role=Role.CLIENT)
        if getattr(request.user, 'centre', None):
            clients = clients.filter(centre=request.user.centre)
        
        serializer = UtilisateurProfilSerializer(clients, many=True)
        return Response(serializer.data)


class AgentDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.role not in [Role.ADMIN, Role.ADMIN_IT]:
            self.permission_denied(request)

    def get_agent(self, agent_id, user):
        """Récupère un agent — admin voit son centre, admin_it voit tout"""
        try:
            filters = {
                'id': agent_id,
            }
            if user.role == Role.ADMIN_IT:
                filters['role__in'] = [Role.AGENT, Role.AGENT_TECHNIQUE, Role.AGENT_ANNEXE, Role.ADMIN]
            else:
                filters['role__in'] = [Role.AGENT, Role.AGENT_TECHNIQUE, Role.AGENT_ANNEXE]
                filters['centre'] = user.centre
            return Utilisateur.objects.get(**filters)
        except Utilisateur.DoesNotExist:
            return None

    def get(self, request, agent_id):
        agent = self.get_agent(agent_id, request.user)
        if not agent:
            return Response({'error': 'Agent introuvable'}, status=status.HTTP_404_NOT_FOUND)
        return Response(UtilisateurProfilSerializer(agent).data)

    def put(self, request, agent_id):
        if request.user.role != Role.ADMIN_IT:
            return Response({'error': 'Seul l\'admin IT peut modifier des agents'}, status=status.HTTP_403_FORBIDDEN)
        agent = self.get_agent(agent_id, request.user)
        if not agent:
            return Response({'error': 'Agent introuvable'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ModifierAgentSerializer(agent, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, agent_id):
        """Supprimer un agent définitivement (admin_it uniquement)"""
        if request.user.role != Role.ADMIN_IT:
            return Response({'error': 'Seul l\'admin IT peut supprimer des agents'}, status=status.HTTP_403_FORBIDDEN)
        agent = self.get_agent(agent_id, request.user)
        if not agent:
            return Response({'error': 'Agent introuvable'}, status=status.HTTP_404_NOT_FOUND)
        from apps.tickets.models import Ticket
        Ticket.objects.filter(agent=agent, statut__in=['soumis', 'en_cours']).update(agent=None, attribution_auto=False)
        agent.delete()
        return Response({'message': 'Agent supprimé avec succès'})


# ============================================================
# HISTORIQUE DES CONNEXIONS (admin seulement)
# ============================================================
class HistoriqueConnexionsView(APIView):
    permission_classes = [IsAuthenticated, EstAdminOuAdminIT]

    def get(self, request):
        """Historique des connexions — admin_it voit TOUS les centres, admin voit son centre"""
        if request.user.role == 'admin':
            users = Utilisateur.objects.filter(centre=request.user.centre)
        else:
            users = Utilisateur.objects.all()
        
        # Filtre par rôle : 'client' ou 'staff' (tout ce qui n'est pas client)
        role_filter = request.query_params.get('role', 'all')
        if role_filter == 'clients':
            users = users.filter(role='client')
        elif role_filter == 'staff':
            users = users.exclude(role='client')
            
        users_ids = users.values_list('id', flat=True)

        historique = HistoriqueConnexion.objects.filter(
            utilisateur_id__in=users_ids
        )

        search = request.query_params.get('search')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if search:
            from django.db.models import Q
            historique = historique.filter(
                Q(utilisateur__nom__icontains=search) | 
                Q(utilisateur__prenom__icontains=search) | 
                Q(utilisateur__email__icontains=search) |
                Q(ip_adresse__icontains=search)
            )
        
        if start_date:
            historique = historique.filter(connecte_a__gte=start_date + 'T00:00:00Z')
        if end_date:
            historique = historique.filter(connecte_a__lte=end_date + 'T23:59:59Z')

        historique = historique.order_by('-connecte_a')[:200]

        serializer = HistoriqueConnexionSerializer(historique, many=True)
        return Response(serializer.data)


# ============================================================
# LIGNES TÉLÉPHONIQUES (client)
# ============================================================
class LignesView(APIView):
    permission_classes = [IsAuthenticated, EstClient]

    def get(self, request):
        """Liste des lignes téléphoniques du client connecté"""
        lignes = LigneTelephonique.objects.filter(client=request.user, actif=True)
        serializer = LigneTelephoniqueSerializer(lignes, many=True)
        return Response(serializer.data)

    def post(self, request):
        """Ajouter une ligne téléphonique"""
        serializer = LigneTelephoniqueSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(client=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LigneDetailView(APIView):
    permission_classes = [IsAuthenticated, EstClient]

    def get_ligne(self, ligne_id, client):
        try:
            return LigneTelephonique.objects.get(id=ligne_id, client=client)
        except LigneTelephonique.DoesNotExist:
            return None

    def get(self, request, ligne_id):
        ligne = self.get_ligne(ligne_id, request.user)
        if not ligne:
            return Response({'error': 'Ligne introuvable'}, status=status.HTTP_404_NOT_FOUND)
        return Response(LigneTelephoniqueSerializer(ligne).data)

    def put(self, request, ligne_id):
        ligne = self.get_ligne(ligne_id, request.user)
        if not ligne:
            return Response({'error': 'Ligne introuvable'}, status=status.HTTP_404_NOT_FOUND)
        serializer = LigneTelephoniqueSerializer(ligne, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, ligne_id):
        ligne = self.get_ligne(ligne_id, request.user)
        if not ligne:
            return Response({'error': 'Ligne introuvable'}, status=status.HTTP_404_NOT_FOUND)
        ligne.actif = False
        ligne.save()
        return Response({'message': 'Ligne désactivée avec succès'})
    
# ============================================================
# HISTORIQUE CONNEXIONS CLIENTS (agent + admin)
# ============================================================
class HistoriqueConnexionsClientsView(APIView):
    permission_classes = [IsAuthenticated, EstAgentOuPlus]

    def get(self, request):
        """Historique des connexions des clients du centre"""
        clients_ids = Utilisateur.objects.filter(
            centre=request.user.centre,
            role='client'
        ).values_list('id', flat=True)

        historique = HistoriqueConnexion.objects.filter(
            utilisateur_id__in=clients_ids
        ).order_by('-connecte_a')[:100]

        serializer = HistoriqueConnexionSerializer(historique, many=True)
        return Response(serializer.data)


# ============================================================
# DEMANDES IT — AGENT
# ============================================================
class DemandesAgentView(APIView):
    permission_classes = [IsAuthenticated, EstAgentOuPlus]

    def get(self, request):
        """Liste des demandes de l'agent connecté"""
        demandes = DemandeIT.objects.filter(demandeur=request.user)
        serializer = DemandeITSerializer(demandes, many=True)
        return Response(serializer.data)

    def post(self, request):
        """Créer une nouvelle demande"""
        serializer = CreerDemandeSerializer(data=request.data)
        if serializer.is_valid():
            # Si c'est un admin, statut directement 'approuvee'
            statut = StatutDemande.APPROUVEE if request.user.role == Role.ADMIN else StatutDemande.EN_ATTENTE
            demande = serializer.save(
                demandeur=request.user,
                centre=request.user.centre,
                statut=statut,
                approuve_par=request.user if request.user.role == Role.ADMIN else None,
            )
            return Response(DemandeITSerializer(demande).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# DEMANDES IT — ADMIN CENTRE
# ============================================================
class DemandesAdminView(APIView):
    permission_classes = [IsAuthenticated, EstAdmin]

    def get(self, request):
        """Demandes des agents de son centre + ses propres demandes"""
        demandes = DemandeIT.objects.filter(centre=request.user.centre)
        serializer = DemandeITSerializer(demandes, many=True)
        return Response(serializer.data)

    def put(self, request, demande_id):
        """Approuver ou refuser une demande"""
        try:
            demande = DemandeIT.objects.get(id=demande_id, centre=request.user.centre)
        except DemandeIT.DoesNotExist:
            return Response({'error': 'Demande introuvable'}, status=status.HTTP_404_NOT_FOUND)

        if demande.statut != StatutDemande.EN_ATTENTE:
            return Response({'error': 'Cette demande a déjà été traitée'}, status=status.HTTP_400_BAD_REQUEST)

        action = request.data.get('action')  # 'approuver' ou 'refuser'
        commentaire = request.data.get('commentaire', '')

        if action == 'approuver':
            demande.statut = StatutDemande.APPROUVEE
            demande.approuve_par = request.user
            demande.reponse_admin = commentaire
            demande.save()
        elif action == 'refuser':
            demande.statut = StatutDemande.REFUSEE
            demande.approuve_par = request.user
            demande.reponse_admin = commentaire
            demande.save()
        else:
            return Response({'error': 'Action invalide (approuver/refuser)'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(DemandeITSerializer(demande).data)


# ============================================================
# DEMANDES IT — ADMIN IT
# ============================================================
class DemandesITView(APIView):
    permission_classes = [IsAuthenticated, EstAdminIT]

    def get(self, request):
        """Toutes les demandes approuvées (de tous les centres)"""
        statut_filter = request.query_params.get('statut', 'approuvee')
        if statut_filter == 'all':
            demandes = DemandeIT.objects.all()
        else:
            demandes = DemandeIT.objects.filter(statut=statut_filter)
        serializer = DemandeITSerializer(demandes, many=True)
        return Response(serializer.data)

    def put(self, request, demande_id):
        """Traiter ou refuser une demande approuvée"""
        try:
            demande = DemandeIT.objects.get(id=demande_id)
        except DemandeIT.DoesNotExist:
            return Response({'error': 'Demande introuvable'}, status=status.HTTP_404_NOT_FOUND)

        if demande.statut != StatutDemande.APPROUVEE:
            return Response({'error': 'Seules les demandes approuvées peuvent être traitées'}, status=status.HTTP_400_BAD_REQUEST)

        action = request.data.get('action')  # 'traiter' ou 'refuser'
        commentaire = request.data.get('commentaire', '')

        if action == 'traiter':
            demande.statut = StatutDemande.TRAITEE
            demande.traite_par = request.user
            demande.reponse_it = commentaire
            # Handle file attachment
            fichier = request.FILES.get('fichier')
            if fichier:
                demande.fichier_reponse_nom = fichier.name
                demande.fichier_reponse_type = fichier.content_type
                demande.fichier_reponse = fichier.read()
            demande.save()
        elif action == 'refuser':
            demande.statut = StatutDemande.REFUSEE_IT
            demande.traite_par = request.user
            demande.reponse_it = commentaire
            demande.save()
        else:
            return Response({'error': 'Action invalide (traiter/refuser)'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(DemandeITSerializer(demande).data)


# ============================================================
# RAPPORT CONNEXIONS (Admin IT)
# ============================================================
class RapportConnexionsView(APIView):
    permission_classes = [IsAuthenticated, EstAdminIT]

    def get(self, request):
        """Génère un rapport XLSX des connexions pour un centre et une période donnée."""
        from io import BytesIO
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        centre_id = request.query_params.get('centre')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if not centre_id or not date_from or not date_to:
            return Response({'detail': 'Les paramètres centre, date_from et date_to sont obligatoires.'}, status=status.HTTP_400_BAD_REQUEST)

        from apps.centres.models import CentreDistribution
        try:
            centre = CentreDistribution.objects.get(id=centre_id)
        except CentreDistribution.DoesNotExist:
            return Response({'detail': 'Centre introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        users_ids = Utilisateur.objects.filter(centre=centre)
        role_filter = request.query_params.get('role')
        if role_filter:
            users_ids = users_ids.filter(role=role_filter)
        users_ids = users_ids.values_list('id', flat=True)
        connexions = HistoriqueConnexion.objects.filter(
            utilisateur_id__in=users_ids,
            connecte_a__gte=date_from + 'T00:00:00Z',
            connecte_a__lte=date_to + 'T23:59:59Z',
        ).order_by('-connecte_a').select_related('utilisateur')

        # Créer le workbook XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Connexions"

        # Styles
        title_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        title_fill = PatternFill(start_color='0055A4', end_color='0055A4', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill(start_color='003D7A', end_color='003D7A', fill_type='solid')
        row_even_fill = PatternFill(start_color='F0F4FA', end_color='F0F4FA', fill_type='solid')
        success_font = Font(name='Calibri', size=10, color='059669')
        fail_font = Font(name='Calibri', size=10, color='DC2626')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Titre
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"Rapport de Connexions — {centre.nom}"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Sous-titre
        ws.merge_cells('A2:I2')
        sub = ws['A2']
        sub.value = f"Période : {date_from} → {date_to}  |  Rôle : {role_filter or 'Tous'}  |  Total : {connexions.count()} connexions"
        sub.font = Font(name='Calibri', italic=True, size=10, color='666666')
        sub.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 22

        # En-têtes
        headers = ['Nom', 'Prénom', 'Email', 'Rôle', 'Adresse IP', 'Navigateur', 'Résultat', 'Connexion', 'Déconnexion']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[4].height = 25

        # Données
        for row_idx, c in enumerate(connexions, 5):
            u = c.utilisateur
            row_data = [
                u.nom, u.prenom, u.email or '',
                u.get_role_display(),
                c.ip_adresse or '', c.user_agent or '',
                'Succès' if c.succes else f'Échec ({c.raison_echec or ""})',
                c.connecte_a.strftime('%d/%m/%Y %H:%M:%S') if c.connecte_a else '',
                c.deconnecte_a.strftime('%d/%m/%Y %H:%M:%S') if c.deconnecte_a else '',
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                if (row_idx - 5) % 2 == 1:
                    cell.fill = row_even_fill
                # Colorer le résultat
                if col_idx == 7:
                    cell.font = success_font if c.succes else fail_font

        # Auto-largeur des colonnes
        col_widths = [15, 15, 25, 16, 16, 30, 20, 20, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"rapport_connexions_{centre.nom.replace(' ', '_')}_{date_from}_{date_to}.xlsx"
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FichierDemandeITView(APIView):
    """Télécharge le fichier joint à une réponse IT."""
    permission_classes = [IsAuthenticated]

    def get(self, request, demande_id):
        try:
            demande = DemandeIT.objects.get(id=demande_id)
        except DemandeIT.DoesNotExist:
            return Response({'error': 'Demande introuvable'}, status=status.HTTP_404_NOT_FOUND)

        if not demande.fichier_reponse:
            return Response({'error': 'Aucun fichier joint'}, status=status.HTTP_404_NOT_FOUND)

        from django.http import HttpResponse as DjHttpResponse
        response = DjHttpResponse(demande.fichier_reponse, content_type=demande.fichier_reponse_type or 'application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{demande.fichier_reponse_nom or "fichier"}"'
        return response